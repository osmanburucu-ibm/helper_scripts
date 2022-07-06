from copy import copy
import xmltodict
from datetime import datetime
from pathlib import Path
import openpyxl
import os, shutil
import logging
import uc_migration_utils as ucutil
import re
import markdownify
from bs4 import BeautifulSoup
from mdutils.mdutils import MdUtils

script_name = "create_missing_blogs"

logging.basicConfig()
logging.root.setLevel(logging.INFO)
logging.basicConfig(format="[%(filename)s:%(lineno)s - %(funcName)20s() ] %(message)s")
logging.basicConfig(level=logging.INFO)

# create file handler which logs even debug messages
fh = logging.FileHandler(f"{script_name}.log", "w+")
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
lformat=logging.Formatter("[%(filename)s:%(lineno)s - %(funcName)20s() ] %(message)s")

fh.setFormatter(lformat)
ch.setFormatter(lformat)
logger1 = logging.getLogger(script_name)


logger1.addHandler(fh)
logger1.addHandler(ch)

IMAGE_URL_RE_PNG='href="\/wp-content\/uploads\/.*?\.png"' # '\/wp-content\/uploads\/.*?\.png'# 'href="\/wp-content\/uploads\/.*?\.png"' # 're.findall(r'"(.*?)"', text1))'
IMAGE_URL_RE_JPG='href="\/wp-content\/uploads\/.*?\.jpg"' # 're.findall(r'"(.*?)"', text1))'
IMAGE_URL_RE_SRC_PNG='src="http:\/\/www\.urbancode\.com\/wp-content\/uploads\/.*?\.png"'
IMAGE_URL_RE_SRC_JPG='src="http:\/\/www\.urbancode\.com\/wp-content\/uploads\/.*?\.jpg"'
IMAGE_RE_SRC_PNG='src=\".*?.png\"'
IMAGE_RE_SRC_JPG='src=\".*?.jpg\"'


CAPTION_RE_START='\[caption id=".*?"\]'
CAPTION_RE_END='\[\/caption\]'

# Create shorthand method for conversion
def md(soup, **options):
    return markdownify.MarkdownConverter(**options).convert_soup(soup)

# create directory with TITLE name
# create first text file (title) which contains the original link!
# search for image links, replace them with local file link, copy image from media directory into local dir
# create a html file with the content, name -> Title
# crate a md file with the content, name -> Title
# 

def copy_image_to_target(targetdir, image_name, blogs_dir, image_path):
    new_image_name = image_name
    if not os.path.exists(f"{targetdir}/{image_name}"):
        if os.path.exists(f"{blogs_dir}/media/{image_path}/{image_name}"): 
            shutil.copyfile(f"{blogs_dir}/media/{image_path}/{image_name}", f"{targetdir}/{image_name}")
        elif x := re.match("-\d*?x\d*?\.png", image_name):
            logger1.info(f"image regex found {image_name}")
            isplit = image_name.split(x.group[0])
            iextension=Path(image_name).suffix
            new_image_name=f"{isplit[0]}.{iextension}"
            logger1.info(f"new_image_name={new_image_name}")
            copy_image_to_target(targetdir, new_image_name, blogs_dir, image_path)
    return new_image_name

def get_publishing_date(ablog):
    
    d = datetime.strptime(ablog["pubDate"], '%a, %d %b %Y %H:%M:%S %z')
    return (d.strftime('%Y.%m.%d'))

def replace_links_in_content (blog_content):
    all_replacable_links = get_list_of_replacable_links()
    new_blog_content = blog_content
    for l in all_replacable_links:
        if l:
            new_blog_content = new_blog_content.replace(l[0], l[1])
    return new_blog_content
    
def create_dir_and_files(ablog):
    config = ucutil.get_config()
    blogs_dir = config[ucutil.BLOGS_DIR]
    targetdir = f"{config[ucutil.BLOGS_DIR]}/migrated/{ablog['title']}"
    ablog_content=ablog['content']
    
    os.makedirs(targetdir, exist_ok=True)
    if not os.path.exists(f"{targetdir}/original_link.txt"):
        with open(f"{targetdir}/original_link.txt", "w") as afile:
            afile.write(ablog['link'])

    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_URL_RE_PNG)
    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_RE_SRC_PNG)
    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_URL_RE_SRC_PNG)    
    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_URL_RE_JPG)
    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_RE_SRC_JPG)
    ablog_content = process_all_images(blogs_dir, targetdir, ablog_content, IMAGE_URL_RE_SRC_JPG)  
             
    d = get_publishing_date(ablog)
    
    ablog_content = replace_links_in_content(ablog_content)
    
    
    new_content = f"<!DOCTYPE html>\n<html><head><title>{ablog['title']}</title></head>\n<body>\n<p><b>This article was originaly published in {d}</b></p>\n{ablog_content}\n</body>\n</html>\n"
    with open(f"{targetdir}/content.html", "w") as afile:
        afile.write(new_content)
    
    md_doc_file = MdUtils(f"{targetdir}/content.md", title=ablog['title'])
    #md_doc_file.new_header(level=1, title=docname)
    soup2 = BeautifulSoup(new_content, "html.parser")
   
    md_doc_file.new_paragraph(md(soup2))
    md_doc_file.create_md_file()

def process_all_images(blogs_dir, targetdir, ablogcontent, regex_image_type):
    ablog_content = ablogcontent
    logger1.info(f"regex_imagetype={regex_image_type}")
    all_images=[]
    r = re.findall(rf"{regex_image_type}", ablog_content) 

    all_images.extend(r)
    logger1.info(f"r={r}")
    for ir in all_images:
        image_with_path=ir.strip()
        image_with_path=image_with_path.replace('http://www.urbancode.com', '')
        image_with_path=image_with_path.replace('href=', '')
        image_with_path=image_with_path.replace('src=', '')
        image_with_path=image_with_path.replace('"', '')
        logger1.info(f"image link={image_with_path}")
        splitted = image_with_path.split("/")
        logger1.info(f"{splitted}")
        image_name = splitted[-1]

        logger1.info(f"Target={targetdir}/{image_name}")
        image_path = Path(image_with_path).parent.absolute()
        logger1.info(f"image_path={image_path}")
        ablog_content=ablog_content.replace(f"{str(image_path)}/", "")
        ablog_content = re.sub(CAPTION_RE_START, "\n", ablog_content, flags=re.MULTILINE)
        ablog_content = re.sub(CAPTION_RE_END, "\n", ablog_content, flags=re.MULTILINE)
       
        new_image_name = copy_image_to_target(targetdir, image_name, blogs_dir, image_path)               
        if (new_image_name != image_name):
            ablog_content=ablog_content.replace(image_name, new_image_name)
    return ablog_content
        
def create_blog(orig_link, my_dict):
    channel = my_dict['rss']['channel']
    ablog = {}
    for i in channel["item"]:
        if (orig_link == i['link']):
            ablog["link"] = i['link']
            ablog["title"] = i['title']
            ablog["pubDate"]=i["pubDate"]
            ablog["content"]=i["content:encoded"]
            logger1.info(f"Title={ablog['title']}")
            create_dir_and_files(ablog)
            
def get_list_of_replacable_links():
    config = ucutil.get_config()
    blogs_dir = config[ucutil.BLOGS_DIR]
    
    list_of_replacable_links=[]
    # load excel with its path
    wrkbk = openpyxl.load_workbook(f"{blogs_dir}/MERGED-all_urls.xlsx")
    sh = wrkbk.active
    
    # iterate through excel and display data
    for i in range(2, sh.max_row+1):
        print("\n")
        print("Row ", i, " data :")
        orig_link=""
        for j in range(1, sh.max_column+1):
            cell_obj = sh.cell(row=i, column=j)
            print (f"{cell_obj.value} - ")
            if j==1: orig_link=cell_obj.value
            if (j==3 and str(cell_obj.value).lower() =="yes"):
                replace_link=cell_obj.value
                break
            else:
                orig_link = ""
                
        if orig_link: 
            logger1.info(f"Original-Link={orig_link} - Replace-Link={replace_link}")
            entry=[orig_link, replace_link]
            list_of_replacable_links.append(entry)
    wrkbk.close()
    return list_of_replacable_links

def main():
    
    config = ucutil.get_config()
    blogs_dir = config[ucutil.BLOGS_DIR]
    with open(f"{blogs_dir}/all_content.urbancode.WordPress.xml", "r") as xml_obj:
        my_dict = xmltodict.parse(xml_obj.read())
        xml_obj.close()
    
    # load excel with its path
    wrkbk = openpyxl.load_workbook(f"{blogs_dir}/MERGED-all_urls.xlsx")
    sh = wrkbk.active
    
    # iterate through excel and display data
    for i in range(2, sh.max_row+1):
        print("\n")
        print("Row ", i, " data :")
        orig_link=""
        for j in range(1, sh.max_column+1):
            cell_obj = sh.cell(row=i, column=j)
            print (f"{cell_obj.value} - ")
            if j==1: orig_link=cell_obj.value
            if ("urbancode" not in orig_link) or (j==2 and cell_obj.value=="YES") or (j==3 and cell_obj.value!=None) or ("www.urbancode.com/plugin" in orig_link):
                orig_link=""
                break
                
        if orig_link: 
            logger1.info(f"Original-Link={orig_link}")
            create_blog(orig_link, my_dict)
            
    wrkbk.close() 
if __name__ == '__main__':
    main()
